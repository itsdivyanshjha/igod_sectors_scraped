"""
Enhanced semantic consolidation analysis for IGOD portals.

This script uses semantic similarity (TF-IDF) instead of keyword matching to identify
REAL consolidation opportunities by detecting functionally similar portals.

Improvements over previous version:
1. Semantic clustering using TF-IDF and cosine similarity
2. Higher similarity thresholds (60%+) for meaningful clusters
3. No double-counting - each portal belongs to only ONE cluster
4. Better organization type detection with hierarchical priority

Usage:
    python sector-wise_checks_semantic.py [input_excel.xlsx] [output_report.xlsx]

Defaults:
    input_excel.xlsx   = "sector-wise_urls.xlsx"
    output_report.xlsx = "igod_sector_consolidation_semantic.xlsx"
"""

import sys
from pathlib import Path
from urllib.parse import urlparse
import re
from collections import defaultdict

import pandas as pd
import numpy as np
from sklearn.feature_extraction.text import TfidfVectorizer
from sklearn.metrics.pairwise import cosine_similarity
from openpyxl.styles import PatternFill


########################
# Pattern definitions #
########################

# Common organizational types (in priority order - most specific first)
ORG_TYPE_PATTERNS = {
    "High Court": r'\bHigh\s+Court\b',
    "District Court": r'\bDistrict\s+Court\b',
    "Supreme Court": r'\bSupreme\s+Court\b',
    "Municipal Corporation": r'\bMunicipal\s+Corporation\b',
    "Development Authority": r'\bDevelopment\s+Authority\b',
    "Police Department": r'\bPolice\b',
    "University": r'\bUniversity\b',
    "College": r'\bCollege\b',
    "Institute": r'\bInstitute\b',
    "Medical College": r'\bMedical\s+College\b',
    "Engineering College": r'\bEngineering\s+College\b',
    "Board": r'\bBoard\b',
    "Directorate": r'\bDirectorate\b',
    "Department": r'\bDepartment\b',
    "Corporation": r'\bCorporation\b',
    "Authority": r'\bAuthority\b',
    "Council": r'\bCouncil\b',
    "Commission": r'\bCommission\b',
    "Tribunal": r'\bTribunal\b',
    "Ministry": r'\bMinistry\b',
}

# Indian states and UTs for location extraction
INDIAN_STATES = [
    "Andhra Pradesh", "Arunachal Pradesh", "Assam", "Bihar", "Chhattisgarh",
    "Goa", "Gujarat", "Haryana", "Himachal Pradesh", "Jharkhand", "Karnataka",
    "Kerala", "Madhya Pradesh", "Maharashtra", "Manipur", "Meghalaya", "Mizoram",
    "Nagaland", "Odisha", "Punjab", "Rajasthan", "Sikkim", "Tamil Nadu",
    "Telangana", "Tripura", "Uttar Pradesh", "Uttarakhand", "West Bengal",
    "Andaman and Nicobar", "Chandigarh", "Dadra and Nagar Haveli", "Daman and Diu",
    "Delhi", "Jammu and Kashmir", "Ladakh", "Lakshadweep", "Puducherry"
]


########################
# Helper functions    #
########################

def extract_state(title: str) -> str:
    """Extract state/UT name from portal title."""
    if not isinstance(title, str):
        return "Central/Unknown"
    
    title_lower = title.lower()
    for state in INDIAN_STATES:
        if state.lower() in title_lower:
            return state
    
    return "Central/Unknown"


def extract_org_type(title: str) -> str:
    """Extract primary organization type from title (hierarchical priority)."""
    if not isinstance(title, str):
        return None
    
    # Check in priority order (more specific first)
    for org_type, pattern in ORG_TYPE_PATTERNS.items():
        if re.search(pattern, title, re.IGNORECASE):
            return org_type
    
    return None


def normalize_for_similarity(title: str) -> str:
    """
    Normalize title for semantic similarity comparison.
    Remove state names but keep functional words.
    """
    if not isinstance(title, str):
        return ""
    
    normalized = title.lower()
    
    # Remove state names
    for state in INDIAN_STATES:
        normalized = re.sub(r'\b' + re.escape(state.lower()) + r'\b', '', normalized)
    
    # Remove common non-functional words
    common_words = ['the', 'of', 'and', '&', 'govt', 'government', 'www', 'http', 'https']
    for word in common_words:
        normalized = re.sub(r'\b' + word + r'\b', '', normalized)
    
    # Clean up special characters but keep important punctuation
    normalized = re.sub(r'[^\w\s\-]', ' ', normalized)
    normalized = re.sub(r'\s+', ' ', normalized).strip()
    
    return normalized


def make_sheet_name(idx: int, name: str, max_len: int = 25) -> str:
    """Create a safe Excel sheet name (max 31 chars)."""
    base = name.replace("&", "and")
    base = re.sub(r"[^0-9A-Za-z \-]", "", base)
    base = re.sub(r"\s+", "_", base.strip())
    base = base[:max_len]
    return f"{idx:02d}_{base}"


##################
# Core functions #
##################

def load_working_df(input_path: Path, sheet_name: str = "Working Portals") -> tuple:
    """Read the working portals sheet and add analysis features.
    
    Returns:
        tuple: (working_df, total_by_sector_dict) - filtered working portals and total counts per sector
    """
    xls = pd.ExcelFile(input_path)
    
    # Handle different column naming conventions
    column_mappings = {
        "Name": "Title",
        "Portal Name": "Title",
        "URL": "Link",
        "Portal URL": "Link",
    }
    
    # Read working portals sheet
    df = pd.read_excel(xls, sheet_name=sheet_name)
    
    # Apply column mappings
    for old_name, new_name in column_mappings.items():
        if old_name in df.columns:
            df = df.rename(columns={old_name: new_name})
    
    # Check for required columns after mapping
    required = ["Sector", "Title", "Link"]
    missing = [c for c in required if c not in df.columns]
    if missing:
        raise ValueError(f"Missing expected columns: {missing}. Available columns: {df.columns.tolist()}")
    
    # Count working portals per sector
    working_counts = df.groupby("Sector").size().to_dict()
    
    # Try to read Errors sheet to get non-working portals count
    error_counts = {}
    if "Errors" in xls.sheet_names:
        try:
            df_errors = pd.read_excel(xls, sheet_name="Errors")
            # Apply same column mappings
            for old_name, new_name in column_mappings.items():
                if old_name in df_errors.columns:
                    df_errors = df_errors.rename(columns={old_name: new_name})
            if "Sector" in df_errors.columns:
                error_counts = df_errors.groupby("Sector").size().to_dict()
        except Exception:
            pass  # If errors sheet can't be read, just use working counts
    
    # Calculate TOTAL portals per sector (working + errors)
    all_sectors = set(working_counts.keys()) | set(error_counts.keys())
    total_by_sector = {}
    for sector in all_sectors:
        total_by_sector[sector] = working_counts.get(sector, 0) + error_counts.get(sector, 0)

    # Filter to only entries with valid URLs
    df = df[df["Link"].notna() & (df["Link"].astype(str).str.strip() != "")].copy()

    # Extract analysis features
    df["State"] = df["Title"].apply(extract_state)
    df["OrgType"] = df["Title"].apply(extract_org_type)
    df["NormalizedTitle"] = df["Title"].apply(normalize_for_similarity)
    df["Domain"] = df["Link"].apply(lambda x: urlparse(x).netloc.lower() if isinstance(x, str) else None)
    
    return df, total_by_sector


def semantic_cluster_by_orgtype(df_subset: pd.DataFrame, org_type: str, 
                                  similarity_threshold: float = 0.60,
                                  min_cluster_size: int = 2) -> list:
    """
    Cluster portals of the same org type using semantic similarity (TF-IDF).
    
    Args:
        df_subset: DataFrame subset for this org type
        org_type: The organization type being clustered
        similarity_threshold: Minimum cosine similarity (0-1) to be in same cluster
        min_cluster_size: Minimum portals required to form a cluster
    
    Returns:
        List of cluster assignments
    """
    if len(df_subset) < min_cluster_size:
        return []
    
    # Get normalized titles
    titles = df_subset["NormalizedTitle"].tolist()
    
    # Skip if all titles are too short or empty
    valid_titles = [t for t in titles if len(t.split()) >= 2]
    if len(valid_titles) < min_cluster_size:
        return []
    
    # Create TF-IDF vectors
    vectorizer = TfidfVectorizer(
        max_features=100,
        stop_words='english',
        ngram_range=(1, 2),  # unigrams and bigrams
        min_df=1
    )
    
    try:
        tfidf_matrix = vectorizer.fit_transform(titles)
    except ValueError:
        # Not enough features
        return []
    
    # Calculate pairwise cosine similarities
    similarity_matrix = cosine_similarity(tfidf_matrix)
    
    # Simple agglomerative clustering based on similarity threshold
    n = len(df_subset)
    clusters = list(range(n))  # Initially each portal is its own cluster
    
    # Merge similar portals into clusters
    for i in range(n):
        if clusters[i] != i:  # Already merged
            continue
        for j in range(i+1, n):
            if similarity_matrix[i][j] >= similarity_threshold:
                # Merge j into i's cluster
                old_cluster = clusters[j]
                for k in range(n):
                    if clusters[k] == old_cluster:
                        clusters[k] = clusters[i]
    
    # Count cluster sizes and filter
    cluster_sizes = {}
    for c in set(clusters):
        cluster_sizes[c] = clusters.count(c)
    
    # Only keep clusters with min_cluster_size or more
    valid_clusters = [c for c in set(clusters) if cluster_sizes[c] >= min_cluster_size]
    
    if not valid_clusters:
        return []
    
    # Build result
    result = []
    for idx, (df_idx, row) in enumerate(df_subset.iterrows()):
        cluster_id = clusters[idx]
        if cluster_id in valid_clusters:
            result.append({
                "df_index": df_idx,
                "cluster_id": cluster_id,
                "cluster_size": cluster_sizes[cluster_id],
                "similarity": similarity_matrix[idx].mean()  # avg similarity to others
            })
    
    return result


def find_cross_sector_patterns(df: pd.DataFrame, min_instances: int = 3) -> pd.DataFrame:
    """
    Find organizational patterns that span across sectors.
    Only includes if there are min_instances or more.
    """
    patterns = []
    
    for org_type in ORG_TYPE_PATTERNS.keys():
        subset = df[df["OrgType"] == org_type].copy()
        
        if len(subset) >= min_instances:
            for idx, row in subset.iterrows():
                patterns.append({
                    "Pattern": org_type,
                    "Total_Instances": len(subset),
                    "Sector": row["Sector"],
                    "State": row["State"],
                    "Title": row["Title"],
                    "Link": row["Link"],
                })
    
    if not patterns:
        return pd.DataFrame()
    
    df_patterns = pd.DataFrame(patterns)
    df_patterns = df_patterns.sort_values(["Pattern", "Sector", "State", "Title"])
    
    return df_patterns


def find_sector_clusters_semantic(df_sector: pd.DataFrame, sector_name: str,
                                   similarity_threshold: float = 0.60,
                                   min_cluster_size: int = 2) -> list:
    """
    Find consolidation clusters within a sector using semantic similarity.
    Each portal belongs to at most ONE cluster (no double-counting).
    
    Args:
        df_sector: DataFrame for this sector
        sector_name: Name of the sector
        similarity_threshold: Cosine similarity threshold (default 0.60 = 60%)
        min_cluster_size: Minimum portals to form a cluster
    
    Returns:
        List of cluster dictionaries
    """
    clusters = []
    clustered_indices = set()  # Track which portals are already clustered
    
    # Group by OrgType and perform semantic clustering
    if "OrgType" not in df_sector.columns:
        return clusters
    
    for org_type in df_sector["OrgType"].dropna().unique():
        subset = df_sector[df_sector["OrgType"] == org_type].copy()
        
        # Skip portals already in a cluster
        subset = subset[~subset.index.isin(clustered_indices)]
        
        if len(subset) < min_cluster_size:
            continue
        
        # Perform semantic clustering
        cluster_assignments = semantic_cluster_by_orgtype(
            subset, org_type, 
            similarity_threshold=similarity_threshold,
            min_cluster_size=min_cluster_size
        )
        
        if not cluster_assignments:
            continue
        
        # Group by cluster_id
        cluster_groups = defaultdict(list)
        for assignment in cluster_assignments:
            cluster_groups[assignment["cluster_id"]].append(assignment)
        
        # Create cluster entries
        for cluster_num, (cluster_id, members) in enumerate(cluster_groups.items()):
            cluster_key = f"{org_type.replace(' ', '_')}_cluster_{cluster_num}"
            
            for member in members:
                df_idx = member["df_index"]
                row = df_sector.loc[df_idx]
                
                clusters.append({
                    "Sector": sector_name,
                    "Cluster": cluster_key,
                    "ClusterType": org_type,
                    "ClusterSize": len(members),
                    "AvgSimilarity": f"{member['similarity']:.2f}",
                    "State": row["State"],
                    "Title": row["Title"],
                    "Link": row["Link"],
                })
                
                # Mark as clustered
                clustered_indices.add(df_idx)
    
    return clusters


def generate_summary(df: pd.DataFrame, sector_clusters: dict, total_by_sector: dict) -> pd.DataFrame:
    """Generate summary sheet showing consolidation opportunities by sector."""
    sectors = sorted(df["Sector"].dropna().unique())
    
    summary_rows = []
    
    for idx, sector in enumerate(sectors, start=1):
        df_sec = df[df["Sector"] == sector]
        working_portals = len(df_sec)
        total_in_sector = total_by_sector.get(sector, working_portals)
        
        # Count clusters in this sector
        clusters = sector_clusters.get(sector, [])
        unique_clusters = len(set(c["Cluster"] for c in clusters))
        portals_in_clusters = len(clusters)
        
        # Calculate consolidation potential
        potential_reduction = portals_in_clusters - unique_clusters if unique_clusters > 0 else 0
        
        sheet_name = make_sheet_name(idx, sector) if unique_clusters > 0 else ""
        
        summary_rows.append({
            "Sector": sector,
            "Sheet": sheet_name,
            "Total Portals in Sector": total_in_sector,
            "Working Portals": working_portals,
            "Consolidation Clusters": unique_clusters,
            "Portals in Clusters": portals_in_clusters,
            "Potential Reduction": potential_reduction,
        })
    
    df_summary = pd.DataFrame(summary_rows)
    
    return df_summary


############
#   Main   #
############

def main():
    # Handle CLI args
    input_path = Path(sys.argv[1]) if len(sys.argv) > 1 else Path("sector-wise-final.xlsx")
    output_path = Path(sys.argv[2]) if len(sys.argv) > 2 else Path("final_sector_consolidation_semantic.xlsx")

    # Configuration
    SIMILARITY_THRESHOLD = 0.60  # 60% semantic similarity required
    MIN_CLUSTER_SIZE = 2  # At least 2 portals to form a cluster
    MIN_CROSS_SECTOR = 3  # At least 3 instances for cross-sector patterns

    print(f"Reading working portals from: {input_path}")
    df, total_by_sector = load_working_df(input_path)
    
    total_all_portals = sum(total_by_sector.values())
    print(f"Loaded {len(df)} working portals across {df['Sector'].nunique()} sectors")
    print(f"Total portals in dataset (working + errors): {total_all_portals}")
    print(f"\nConfiguration:")
    print(f"  • Similarity threshold: {SIMILARITY_THRESHOLD*100:.0f}%")
    print(f"  • Minimum cluster size: {MIN_CLUSTER_SIZE}")
    print(f"  • No double-counting: Each portal in max 1 cluster")
    
    # ===== Step 1: Find cross-sector patterns =====
    print("\n" + "="*60)
    print("Analyzing cross-sector patterns...")
    print("="*60)
    df_patterns = find_cross_sector_patterns(df, min_instances=MIN_CROSS_SECTOR)
    
    if len(df_patterns) > 0:
        pattern_counts = df_patterns.groupby("Pattern")["Total_Instances"].first().to_dict()
        print(f"Found {len(pattern_counts)} cross-sector patterns:")
        for pattern, count in sorted(pattern_counts.items(), key=lambda x: -x[1])[:15]:
            print(f"  • {pattern}: {count} instances")
    else:
        print("No cross-sector patterns found.")
    
    # ===== Step 2: Find sector-specific clusters with semantic similarity =====
    print("\n" + "="*60)
    print("Analyzing sector-specific consolidation (semantic clustering)...")
    print("="*60)
    sectors = sorted(df["Sector"].dropna().unique())
    sector_clusters = {}
    sector_sheets = {}
    
    for idx, sector in enumerate(sectors, start=1):
        df_sec = df[df["Sector"] == sector].copy()
        clusters = find_sector_clusters_semantic(
            df_sec, sector,
            similarity_threshold=SIMILARITY_THRESHOLD,
            min_cluster_size=MIN_CLUSTER_SIZE
        )
        
        if clusters:
            sector_clusters[sector] = clusters
            
            # Create sheet
            df_cluster = pd.DataFrame(clusters)
            df_cluster = df_cluster.sort_values(["Cluster", "State", "Title"])
            
            sheet_name = make_sheet_name(idx, sector)
            sector_sheets[sheet_name] = df_cluster
            
            unique_clusters = len(set(c["Cluster"] for c in clusters))
            print(f"  {sector}: {unique_clusters} clusters, {len(clusters)} portals")
    
    # ===== Step 3: Generate summary =====
    print("\n" + "="*60)
    print("Generating summary...")
    print("="*60)
    df_summary = generate_summary(df, sector_clusters, total_by_sector)
    
    total_potential = df_summary["Potential Reduction"].sum()
    total_working_portals = df_summary["Working Portals"].sum()
    total_all_portals = df_summary["Total Portals in Sector"].sum()
    print(f"  Total portals: {total_all_portals} (Working: {total_working_portals})")
    print(f"  Total potential portal reduction: {total_potential} portals")
    
    # ===== Step 4: Create consolidated "All_Clusters" sheet =====
    print("\nCreating consolidated clusters sheet...")
    all_clusters_data = []
    for sector, clusters in sector_clusters.items():
        all_clusters_data.extend(clusters)
    
    if all_clusters_data:
        df_all_clusters = pd.DataFrame(all_clusters_data)
        df_all_clusters = df_all_clusters.sort_values(["Sector", "Cluster", "State", "Title"])
    else:
        df_all_clusters = pd.DataFrame()
    
    # ===== Step 5: Write output =====
    print(f"\nWriting report to: {output_path}")
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        # Summary sheet
        df_summary.to_excel(writer, sheet_name="Summary", index=False)
        
        # Consolidated all clusters sheet (plain)
        if len(df_all_clusters) > 0:
            df_all_clusters.to_excel(writer, sheet_name="All_Clusters", index=False)
            
            # Color-coded version of All_Clusters
            df_all_clusters.to_excel(writer, sheet_name="All_Clusters_Colored", index=False)
        
        # Cross-sector patterns sheet
        if len(df_patterns) > 0:
            df_patterns.to_excel(writer, sheet_name="All_Patterns", index=False)
        
        # Sector-specific sheets
        for sheet_name, df_sheet in sector_sheets.items():
            df_sheet.to_excel(writer, sheet_name=sheet_name, index=False)
    
    # ===== Step 6: Apply color coding to All_Clusters_Colored sheet =====
    if len(df_all_clusters) > 0:
        print("Applying color coding to All_Clusters_Colored sheet...")
        apply_cluster_colors(output_path, df_all_clusters)
    
    # ===== Step 7: Print final summary =====
    print_final_summary(df, sectors, df_patterns, sector_sheets, total_potential, 
                        total_working_portals, total_all_portals, output_path, SIMILARITY_THRESHOLD)


def apply_cluster_colors(output_path: Path, df_clusters: pd.DataFrame):
    """Apply pastel color coding to each cluster in the All_Clusters_Colored sheet."""
    from openpyxl import load_workbook
    
    # Pastel colors for clusters (easily readable)
    PASTEL_COLORS = [
        "FFE6E6",  # Light pink
        "E6FFE6",  # Light green
        "E6E6FF",  # Light blue
        "FFFFE6",  # Light yellow
        "FFE6FF",  # Light magenta
        "E6FFFF",  # Light cyan
        "FFF0E6",  # Light peach
        "F0E6FF",  # Light lavender
        "E6FFF0",  # Light mint
        "FFF5E6",  # Light apricot
        "F5E6FF",  # Light violet
        "E6F5FF",  # Light sky blue
        "FFE6F0",  # Light rose
        "E6FFE0",  # Light lime
        "E0E6FF",  # Light periwinkle
        "FFFAE6",  # Light cream
        "FAE6FF",  # Light orchid
        "E6FAFF",  # Light aqua
        "FFE6E0",  # Light coral
        "E0FFE6",  # Light seafoam
        "FFE0E6",  # Light blush
        "E6E0FF",  # Light iris
        "E0FFFA",  # Light turquoise
        "FFEDE6",  # Light salmon
        "EDE6FF",  # Light wisteria
    ]
    
    # Load workbook
    wb = load_workbook(output_path)
    ws = wb["All_Clusters_Colored"]
    
    # Get unique clusters and assign colors
    unique_clusters = df_clusters["Cluster"].unique().tolist()
    cluster_colors = {}
    for i, cluster in enumerate(unique_clusters):
        color_idx = i % len(PASTEL_COLORS)
        cluster_colors[cluster] = PASTEL_COLORS[color_idx]
    
    # Apply colors to rows (skip header row)
    for row_idx, cluster_name in enumerate(df_clusters["Cluster"].tolist(), start=2):
        color = cluster_colors.get(cluster_name, "FFFFFF")
        fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
        
        # Apply to all cells in the row
        for col_idx in range(1, len(df_clusters.columns) + 1):
            ws.cell(row=row_idx, column=col_idx).fill = fill
    
    # Save workbook
    wb.save(output_path)
    print(f"  Applied {len(unique_clusters)} unique cluster colors")


def print_final_summary(df, sectors, df_patterns, sector_sheets, total_potential, total_working_portals, total_all_portals, output_path, similarity_threshold):
    """Print the final summary after all processing is complete."""
    print("\n" + "="*60)
    print("✅ SEMANTIC CONSOLIDATION ANALYSIS COMPLETE")
    print("="*60)
    print(f"\n📊 Results:")
    print(f"  • Total portals in dataset: {total_all_portals}")
    print(f"  • Working portals analyzed: {total_working_portals}")
    print(f"  • Sectors analyzed: {len(sectors)}")
    print(f"  • Cross-sector patterns: {len(df_patterns) if len(df_patterns) > 0 else 0}")
    print(f"  • Sectors with consolidation opportunities: {len(sector_sheets)}")
    print(f"  • Potential portal reduction: {total_potential} portals")
    print(f"\n📁 Output file: {output_path}")
    print(f"\n📋 Features:")
    print(f"  ✓ Semantic clustering (TF-IDF) with {similarity_threshold*100:.0f}% similarity threshold")
    print(f"  ✓ No double-counting - each portal in exactly 1 cluster")
    print(f"  ✓ AvgSimilarity column shows cluster quality")
    print(f"  ✓ Color-coded All_Clusters_Colored sheet for easy visualization")
    print(f"\n💡 Next steps:")
    print(f"  1. Open 'Summary' sheet for sector-wise overview")
    print(f"  2. Share 'All_Clusters_Colored' sheet with your mentor (color-coded!)")
    print(f"  3. Check 'AvgSimilarity' values - higher = better cluster quality")


if __name__ == "__main__":
    main()

